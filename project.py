import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook

class TodoListApp:
    def __init__(self, master):
        self.master = master
        self.master.title("To-Do List App")
        self.tasks = []
        self.master.config(bg="#F0F0F0") 
        
        self.task_label = tk.Label(master, text="To-Do List", bg="#F0F0F0", fg="#333333", font=("Times New Roman", 20, "bold"))
        self.task_label.grid(row=0, column=0, padx=5, pady=(2,100))

        self.user_label = tk.Label(master, text="User Name:", bg="#F0F0F0", fg="#333333", font=("Times New Roman", 15, "bold"))
        self.user_label.grid(row=1, column=0, padx=5, pady=(2,70), sticky="w")

        self.user_entry = tk.Entry(master, width=50)
        self.user_entry.grid(row=1, column=1, padx=10, pady=(2,70))

        self.task_label = tk.Label(master, text="Enter The Task:", bg="#F0F0F0", fg="#333333", font=("Times New Roman", 15, "bold"))
        self.task_label.grid(row=2, column=0, padx=5, pady=10, sticky="w")

        self.date_label = tk.Label(master, text="Deadline Date:", bg="#F0F0F0", fg="#333333", font=("Times New Roman", 15, "bold"))
        self.date_label.grid(row=3, column=0, padx=5, pady=10, sticky="w")

        self.time_label = tk.Label(master, text="Deadline Time:", bg="#F0F0F0", fg="#333333", font=("Times New Roman", 15, "bold"))
        self.time_label.grid(row=4, column=0, padx=5, pady=10, sticky="w")

        self.task_entry = tk.Entry(master, width=50)
        self.task_entry.grid(row=2, column=1, padx=10, pady=10)

        self.date_entry = tk.Entry(master, width=50,)
        self.date_entry.grid(row=3, column=1, padx=10, pady=10)

        self.time_entry = tk.Entry(master, width=50)
        self.time_entry.grid(row=4, column=1, padx=10, pady=10)

        self.task_listbox_label = tk.Label(master, text="Tasks:", bg="#F0F0F0", fg="#333333", font=("Times New Roman", 15, "bold"))
        self.task_listbox_label.grid(row=5, column=0, padx=5, pady=5, sticky="w")

        self.task_listbox = tk.Listbox(master, width=50, height=15)
        self.task_listbox.grid(row=6, column=1, padx=5, pady=5)

        self.load_button = tk.Button(master, text="Load A Task", command=self.load_tasks, bg="yellow", fg="black", font=("Times New Roman", 15, "bold"))
        self.load_button.grid(row=1, column=2, padx=5, pady=(2,70), sticky="e")

        self.add_button = tk.Button(master, text="Add A Task", command=self.add_task, bg="#4CAF50", fg="white", font=("Times New Roman", 15, "bold"))
        self.add_button.grid(row=3, column=2, padx=5, pady=5, sticky="e")

        self.delete_button = tk.Button(master, text="Delete A Task", command=self.delete_task, bg="#E72929", fg="white", font=("Times New Roman", 15, "bold"))
        self.delete_button.grid(row=6, column=2, padx=5, pady=5, sticky="w")

        self.save_button = tk.Button(master, text="Save Tasks", command=self.save_tasks, bg="#3498DB", fg="black", font=("Times New Roman", 15, "bold"))
        self.save_button.grid(row=7, column=1, padx=(2,100), pady=20, sticky="e")

        self.new_task_button = tk.Button(master, text="New Task", command=self.new_task_interface, bg="#FF5733", fg="white", font=("Times New Roman", 18, "bold"))
        self.new_task_button.grid(row=0, column=3, padx=(2,10), pady=5, sticky="e")

    def add_task(self):
        task = self.task_entry.get().strip()
        time = self.time_entry.get().strip()
        date = self.date_entry.get().strip()
        if task and time and date:
            task_with_time = f"{task} - {time}"
            self.tasks.append((len(self.tasks) + 1, task, time, date))
            self.update_task_listbox()

    def delete_task(self):
        selected_index = self.task_listbox.curselection()
        if selected_index:
            index = selected_index[0]
            del self.tasks[index]
            self.update_task_listbox()

    def update_task_listbox(self):
        self.task_listbox.delete(0, tk.END)
        for task_num, task, time, date in self.tasks:
            self.task_listbox.insert(tk.END, f"{task_num}. {task} - {time} - {date}")

    def save_tasks(self):
        if not self.tasks:
            messagebox.showinfo("Info", "No tasks to save!")
            return

        user_name = self.user_entry.get().strip()
        if not user_name:
            messagebox.showerror("Error", "Please enter a user name!")
            return

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Tasks"
            sheet.append(["Task Number", "Task", "Time", "Date"])
            for task_num, task, time, date in self.tasks:
                sheet.append([task_num, task, time, date])
            workbook.save(f"{user_name}_tasks.xlsx")
            messagebox.showinfo("Success", "Tasks saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save tasks: {str(e)}")

    def new_task_interface(self):
        self.task_entry.delete(0, tk.END)
        self.time_entry.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)
        self.tasks.clear()
        self.update_task_listbox()

    def load_tasks(self):
        user_name = self.user_entry.get().strip()
        if not user_name:
            messagebox.showerror("Error", "Please enter a user name!")
            return

        try:
            workbook = openpyxl.load_workbook(f"{user_name}_tasks.xlsx")
            sheet = workbook["Tasks"]
            self.tasks.clear()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.tasks.append(row)
            self.update_task_listbox()
            messagebox.showinfo("Success", "Tasks loaded successfully!")
        except FileNotFoundError:
            messagebox

def main():
    root = tk.Tk()
    app = TodoListApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()